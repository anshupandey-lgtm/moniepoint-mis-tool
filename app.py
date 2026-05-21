import json
import re
from io import BytesIO
from pathlib import Path
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Moniepoint MIS Automation", layout="wide")

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
MAPPING_FILE = DATA_DIR / "mapping_master.json"
TEMPLATE_FILE = APP_DIR / "Moniepoint-India-Management-report-Apr-2026.xlsx"


def load_mapping_master():
    if MAPPING_FILE.exists():
        return json.loads(MAPPING_FILE.read_text())
    seed = {
        "SALES": {"account_type": "Income", "head": "Revenue from Operations"},
        "SALARIES & WAGES": {"account_type": "Expense", "head": "Salaries & wages"},
        "PERFORMANCE BONUSES AND INCENTIVES": {"account_type": "Expense", "head": "Performance bonuses"},
        "LEAVE ENCASHMENT": {"account_type": "Expense", "head": "Other employee costs", "sub_head": "Leave Encashment"},
        "PF ADMIN CHARGES": {"account_type": "Expense", "head": "Other employee costs", "sub_head": "PF Admin Charges"},
        "EMPLOYER PF CONTRIBUTION": {"account_type": "Expense", "head": "Other employee costs", "sub_head": "Employers Contribution to Social Security Funds"},
        "EMPLOYEE BENEFITS (PRIVATE MEDICAL INSURANCE)": {"account_type": "Expense", "head": "Insurance"},
        "ACCOUNTING AND TAX CONSULTING FEES": {"account_type": "Expense", "head": "Legal and professional costs", "sub_head": "Accounting & Bookkeeping"},
        "DEPRECIATION": {"account_type": "Expense", "head": "Depreciation"},
        "RENT AND RATES": {"account_type": "Expense", "head": "Rent and other office expenses"},
        "EXCHANGE GAIN OR LOSS": {"account_type": "Expense", "head": "Unrealised FX gains or losses"},
        "CORPORATE INCOME TAX": {"account_type": "Expense", "head": "Income tax"},
        "ACCOUNTS PAYABLE": {"account_type": "Liability", "head": "Trade Payables"},
        "ACCRUED LEGAL AND PROFESSIONAL EXPENSES": {"account_type": "Liability", "head": "Accrued Legal and Professional Expenses"},
        "COMPUTERS AND LAPTOPS": {"account_type": "Asset", "head": "Property, Plant & Equipment and Intangible Assets", "sub_head": "COMPUTERS AND LAPTOPS"},
        "PREPAID EXPENSES": {"account_type": "Asset", "head": "Prepaid Expenses"},
        "ACCOUNTS RECEIVABLE": {"account_type": "Asset", "head": "Trade Receivables"},
        "RAZORPAY WALLET": {"account_type": "Asset", "head": "Other Current Assets", "sub_head": "Razorpay Wallet"},
        "TDS RECEIVABLE": {"account_type": "Asset", "head": "Other Current Assets", "sub_head": "TDS Receivable"},
        "ADVANCE TAX": {"account_type": "Asset", "head": "Other Current Assets", "sub_head": "Advance Tax"},
        "INTERCOMPANY RECEIVABLES - MONIEPOINT UK": {"account_type": "Asset", "head": "Other Assets", "sub_head": "Intercompany Receivables - Moniepoint UK"},
        "INTERCOMPANY RECEIVABLES - MONIEPOINT INC": {"account_type": "Asset", "head": "Other Assets", "sub_head": "Intercompany Receivables - Moniepoint Inc"},
        "INTERCOMPANY PAYABLES - MONIEPOINT UK": {"account_type": "Liability", "head": "Other Current Liabilities", "sub_head": "Intercompany Payables - Moniepoint UK"},
        "INTERCOMPANY PAYABLES - MONIEPOINT INC": {"account_type": "Liability", "head": "Other Current Liabilities", "sub_head": "Intercompany Payables - Moniepoint Inc"},
        "ACCRUED PAYROLL EXPENSE": {"account_type": "Liability", "head": "Accrued expenses and other payables", "sub_head": "Accrued Payroll Expense"},
        "ACCUMULATED DEPRECIATION - COMPUTERS": {"account_type": "Asset", "head": "Property, Plant & Equipment and Intangible Assets", "sub_head": "Accumulated Depreciation - Computers"},
        "DEFERRED TAX ASSET": {"account_type": "Asset", "head": "Other Current Assets", "sub_head": "Deferred Tax Asset"},
        "CORPORATE INCOME TAX PAYABLE": {"account_type": "Liability", "head": "Short Term Provisions", "sub_head": "Corporate Income Tax Payable"},
        "PROVISION FOR PROFESSIONAL TAX": {"account_type": "Liability", "head": "Short Term Provisions", "sub_head": "Provision for Professional Tax"},
    }
    MAPPING_FILE.write_text(json.dumps(seed, indent=2))
    return seed


def normalize_name(v):
    return re.sub(r"\s+", " ", str(v).strip().upper())


def read_tb(upload):
    xls = pd.ExcelFile(upload)
    sheets = xls.sheet_names
    bs_sheet = next((s for s in sheets if any(k in s.lower() for k in ["bs", "balance"])), sheets[0])
    pl_sheet = next((s for s in sheets if any(k in s.lower() for k in ["p&l", "pl", "profit"])), sheets[1] if len(sheets) > 1 else sheets[0])
    return pd.read_excel(xls, sheet_name=bs_sheet), pd.read_excel(xls, sheet_name=pl_sheet)


def read_salary(upload):
    xls = pd.ExcelFile(upload)
    return pd.read_excel(xls, sheet_name=xls.sheet_names[0])


def read_reference(upload):
    xls = pd.ExcelFile(upload)
    return pd.read_excel(xls, sheet_name=xls.sheet_names[0])


def extract_rate(ref_df):
    cols = {str(c).strip().lower(): c for c in ref_df.columns}
    pair_col = next((cols[k] for k in cols if "currency" in k or "pair" in k), ref_df.columns[0])
    rate_col = next((cols[k] for k in cols if "rate" in k), ref_df.columns[min(1, len(ref_df.columns)-1)])
    temp = ref_df.copy()
    temp = temp[temp[pair_col].astype(str).str.contains("USD", case=False, na=False)]
    return round(pd.to_numeric(temp[rate_col], errors="coerce").dropna().mean(), 4)


def salary_summary(df):
    cols = [str(c).strip() for c in df.columns]
    canon = {c.upper(): c for c in cols}
    gross_col = canon.get("GROSS SALARY") or canon.get("GROSS PAY") or canon.get("GROSS")
    def n(c):
        if c in df.columns:
            return pd.to_numeric(df[c], errors="coerce").fillna(0)
        return pd.Series([0] * len(df), index=df.index)
    gross = n(gross_col) if gross_col else pd.Series([0]*len(df), index=df.index)
    basic = n(canon.get("BASIC SALARY", "Basic Salary"))
    da = n(canon.get("DA", "DA"))
    return {
        "rows": int(len(df)),
        "gross_total": float(gross.sum()),
        "basic_da_total": float((basic + da).sum()),
        "pf_admin_0_5pct_basic_da": float(round(((basic + da).sum()) * 0.005, 0)),
    }


def ledger_amount_frame(df):
    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]
    ledger_col = next((c for c in work.columns if "ledger" in c.lower() or "account" in c.lower()), work.columns[0])
    amount_col = next((c for c in work.columns if "net" in c.lower() or "amount" in c.lower() or "debit" in c.lower()), work.columns[-1])
    out = work[[ledger_col, amount_col]].copy()
    out.columns = ["ledger", "amount"]
    out["ledger"] = out["ledger"].astype(str).map(normalize_name)
    out["amount"] = pd.to_numeric(out["amount"], errors="coerce").fillna(0)
    return out.groupby("ledger", as_index=False)["amount"].sum()


def match_mapping(ledger, mapping):
    key = normalize_name(ledger)
    if key in mapping:
        return key, mapping[key]
    for k, v in mapping.items():
        if k in key or key in k:
            return k, v
    return None, None


def classify(df, mapping):
    out = df.copy()
    out["mapping_key"] = ""
    out["account_type"] = ""
    out["head"] = ""
    out["sub_head"] = ""
    out["status"] = "Mapped"
    for i, row in out.iterrows():
        key, info = match_mapping(row["ledger"], mapping)
        if info:
            out.at[i, "mapping_key"] = key
            out.at[i, "account_type"] = info.get("account_type", "")
            out.at[i, "head"] = info.get("head", "")
            out.at[i, "sub_head"] = info.get("sub_head", "")
        else:
            out.at[i, "status"] = "Unmapped"
    return out


def _copy_sheet(src, dst):
    for row in src.iter_rows():
        for cell in row:
            nc = dst[cell.coordinate]
            nc.value = cell.value
            if cell.has_style:
                nc._style = copy(cell._style)
            if cell.number_format:
                nc.number_format = cell.number_format
            if cell.font:
                nc.font = copy(cell.font)
            if cell.fill:
                nc.fill = copy(cell.fill)
            if cell.border:
                nc.border = copy(cell.border)
            if cell.alignment:
                nc.alignment = copy(cell.alignment)
            if cell.protection:
                nc.protection = copy(cell.protection)
    for key, dim in src.column_dimensions.items():
        dst.column_dimensions[key].width = dim.width
        dst.column_dimensions[key].hidden = dim.hidden
    for key, dim in src.row_dimensions.items():
        dst.row_dimensions[key].height = dim.height
        dst.row_dimensions[key].hidden = dim.hidden
    for merged in src.merged_cells.ranges:
        dst.merge_cells(str(merged))


def build_output(template_bytes, tb_bs, tb_pl, salary_df, ref_df):
    wb = load_workbook(BytesIO(template_bytes))
    rate = extract_rate(ref_df)
    summary = salary_summary(salary_df)
    bs_mapped = classify(ledger_amount_frame(tb_bs), load_mapping_master()) if not tb_bs.empty else pd.DataFrame()
    pl_mapped = classify(ledger_amount_frame(tb_pl), load_mapping_master()) if not tb_pl.empty else pd.DataFrame()
    for name in ["BS INR", "PL INR", "BS USD", "PL USD", "Apr 26 Reference Rates", "Apr 26 Salary Register"]:
        if name not in wb.sheetnames:
            continue
    if "Apr 26 Reference Rates" in wb.sheetnames:
        ws = wb["Apr 26 Reference Rates"]
        for i, col in enumerate(ref_df.columns, start=1):
            ws.cell(row=4, column=i).value = col
        for r, row in enumerate(ref_df.itertuples(index=False), start=5):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c).value = v
    if "Apr 26 Salary Register" in wb.sheetnames:
        ws = wb["Apr 26 Salary Register"]
        for i, col in enumerate(salary_df.columns, start=1):
            ws.cell(row=4, column=i).value = col
        for r, row in enumerate(salary_df.itertuples(index=False), start=5):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c).value = v
    if "BS INR" in wb.sheetnames and not bs_mapped.empty:
        ws = wb["BS INR"]
        for idx, row in bs_mapped.iterrows():
            ws.cell(row=10 + idx, column=1).value = row["ledger"]
            ws.cell(row=10 + idx, column=2).value = row["amount"]
            ws.cell(row=10 + idx, column=3).value = row["mapping_key"]
            ws.cell(row=10 + idx, column=4).value = row["account_type"]
            ws.cell(row=10 + idx, column=5).value = row["head"]
            ws.cell(row=10 + idx, column=6).value = row["sub_head"]
    if "PL INR" in wb.sheetnames and not pl_mapped.empty:
        ws = wb["PL INR"]
        for idx, row in pl_mapped.iterrows():
            ws.cell(row=10 + idx, column=1).value = row["ledger"]
            ws.cell(row=10 + idx, column=2).value = row["amount"]
            ws.cell(row=10 + idx, column=3).value = row["mapping_key"]
            ws.cell(row=10 + idx, column=4).value = row["account_type"]
            ws.cell(row=10 + idx, column=5).value = row["head"]
            ws.cell(row=10 + idx, column=6).value = row["sub_head"]
    if "BS USD" in wb.sheetnames and not bs_mapped.empty:
        ws = wb["BS USD"]
        for idx, row in bs_mapped.iterrows():
            ws.cell(row=10 + idx, column=1).value = row["ledger"]
            ws.cell(row=10 + idx, column=2).value = row["amount"] / rate if rate else row["amount"]
    if "PL USD" in wb.sheetnames and not pl_mapped.empty:
        ws = wb["PL USD"]
        for idx, row in pl_mapped.iterrows():
            ws.cell(row=10 + idx, column=1).value = row["ledger"]
            ws.cell(row=10 + idx, column=2).value = row["amount"] / rate if rate else row["amount"]
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out, summary, rate


def main():
    st.title("Moniepoint MIS Automation")
    tb_file = st.file_uploader("Upload TB file", type=["xlsx"])
    salary_file = st.file_uploader("Upload Salary Register", type=["xlsx"])
    ref_file = st.file_uploader("Upload Reference Rates", type=["xlsx"])
    template_file = st.file_uploader("Upload previous month MIS template", type=["xlsx"])
    if tb_file and salary_file and ref_file and template_file:
        tb_bs, tb_pl = read_tb(tb_file)
        salary_df = read_salary(salary_file)
        ref_df = read_reference(ref_file)
        template_bytes = template_file.getvalue()
        output, summary, rate = build_output(template_bytes, tb_bs, tb_pl, salary_df, ref_df)
        st.subheader("Summary")
        st.json(summary)
        st.write("USD rate:", rate)
        st.download_button("Download extended MIS", data=output, file_name="Moniepoint_India_Management_report_extended.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Upload TB, Salary Register, Reference Rates, and previous MIS template.")

if __name__ == "__main__":
    main()
